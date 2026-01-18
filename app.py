
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
import copy
from datetime import datetime
from dateutil.relativedelta import relativedelta
import re
import os
import io
import random # Import random for colors

# Define colors
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Define thin border style
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# ------------------------------------------------
# Συναρτήσεις ημερομηνιών
# ------------------------------------------------
def parse_date(text, is_start=True):
    text = str(text).strip()
    if "σήμερα" in text.lower() or "simera" in text.lower():
        if not is_start:
            return datetime.today()
        else:
            pass # This case is tricky for start date without end date context

    if re.match(r"^\d{4}$", text):
        if is_start:
            return datetime.strptime("01/01/" + text, "%d/%m/%Y")
        else:
            return datetime.strptime("31/12/" + text, "%d/%m/%Y")
    elif re.match(r"^\d{1,2}/\d{4}$", text):
        if is_start:
            return datetime.strptime("01/" + text, "%d/%m/%Y")
        else:
            d = datetime.strptime("01/" + text, "%d/%m/%Y")
            return d + relativedelta(months=1) - relativedelta(days=1)
    elif re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", text):
        return datetime.strptime(text, "%d/%m/%Y")
    else:
        raise ValueError(f"Unsupported date format: '{text}'. Expected 'YYYY', 'M/YYYY' or 'MM/YYYY', 'D/M/YYYY' or 'DD/MM/YYYY', or 'Σήμερα' (for end date).")

def parse_period(p):
    p_cleaned = str(p).strip().replace("—", "-").replace("–", "-")
    if re.match(r"^\d{4}$", p_cleaned):
        return parse_date(p_cleaned, True), parse_date(p_cleaned, False)

    parts = p_cleaned.split("-")
    if len(parts) != 2:
        raise ValueError(f"Invalid period format: '{p}'. Expected 'YYYY' or 'START_DATE-END_DATE'.")
    a, b = parts
    return parse_date(a, True), parse_date(b, False)

def month_range(start, end):
    current = datetime(start.year, start.month, 1)
    end = datetime(end.year, end.month, 1)
    out = []
    while current <= end:
        out.append((current.year, current.month))
        current += relativedelta(months=1)
    return out

def is_light_color(hex_color):
    hex_color = hex_color.lstrip('#')
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    luminance = (0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]) / 255
    return luminance > 0.5

def process_excel_data(uploaded_template_file, uploaded_input_file):
    # ------------------------------------------------
    # Read INPUT and find columns
    # ------------------------------------------------
    wb_in = openpyxl.load_workbook(uploaded_input_file)
    ws_in = wb_in.active

    headers = {}
    for c in range(1, ws_in.max_column + 1):
        val = str(ws_in.cell(1,c).value).strip()
        headers[val] = c

    if "ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ" not in headers or "ΑΝΘΡΩΠΟΜΗΝΕΣ" not in headers:
        st.error("Το input πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ")
        return None, None, None, None, None # Indicate error

    PERIOD_COL = headers["ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ"]
    AM_COL = headers["ΑΝΘΡΩΠΟΜΗΝΕΣ"] # Fixed typo here

    data = []
    all_months = set()
    project_counter = 0 # Initialize project counter

    yellow_fill_rgb_check = "FFFF00"

    total_all_projects_am = 0

    for r in range(2, ws_in.max_row + 1):
        period_cell = ws_in.cell(r, PERIOD_COL)
        period = period_cell.value
        am_raw = ws_in.cell(r, AM_COL).value
        try:
            am = int(am_raw) if am_raw is not None else 0
        except (ValueError, TypeError):
            am = 0

        if not period or am == 0:
            continue

        try:
            start, end = parse_period(str(period))
        except ValueError as e:
            st.warning(f"Skipping row {r} due to period parsing error: {e}")
            continue

        months = month_range(start, end)
        months_in_period_count = len(months)

        if months_in_period_count > 0:
            am_per_month_ratio = am / months_in_period_count
        else:
            am_per_month_ratio = 0

        cell_rgb = period_cell.fill.start_color.rgb if period_cell.fill.start_color else None
        is_yellow = (cell_rgb == yellow_fill_rgb_check or cell_rgb == "FFFFFF00")

        data.append({
            "project_id": project_counter,
            "period_str": period,
            "original_am": am,
            "months_in_period": months,
            "months_in_period_count": months_in_period_count,
            "am_per_month_ratio": am_per_month_ratio,
            "allocated_am": 0,
            "unallocated_am": am,
            "is_yellow": is_yellow,
            "excel_row": 0,
            "reasons_log": []
        })
        project_counter += 1

        for m in months:
            all_months.add(m)

        total_all_projects_am += am

    all_months = sorted(all_months)
    years = sorted(set(y for y,m in all_months))

    data.sort(key=lambda x: (not x["is_yellow"], x["months_in_period_count"])) # MODIFIED LINE

    project_id_map = {proj['project_id']: proj for proj in data}

    # ------------------------------------------------
    # Open TEMPLATE
    # ------------------------------------------------
    wb = openpyxl.load_workbook(uploaded_template_file)
    ws = wb.active # This is the sheet that will become 'ΑΝΑΛΥΣΗ'

    ws.freeze_panes = 'D1'

    START_ROW_DATA = 4
    YEAR_ROW = 2
    MONTH_ROW = 3
    YEARLY_TOTAL_ROW = START_ROW_DATA + 1
    START_COL = 5
    MAX_YEARLY_CAPACITY = 11

    ws.title = 'ΑΝΑΛΥΣΗ'

    cv_sheet = wb.create_sheet(title='CV', index=0)
    for row_idx, row_data in enumerate(ws_in.iter_rows()):
        for col_idx, cell in enumerate(row_data):
            new_cell = cv_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format

    for col_idx in range(1, ws_in.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in ws_in.column_dimensions:
            cv_sheet.column_dimensions[col_letter].width = ws_in.column_dimensions[col_letter].width

    cv_sheet['A1'] = 'Α/Α'
    cv_sheet['A1'].font = Font(bold=True)
    cv_sheet['A1'].border = thin_border

    last_row_b = 0
    for row_idx_cv in range(1, cv_sheet.max_row + 1):
        if cv_sheet.cell(row=row_idx_cv, column=2).value is not None:
            last_row_b = row_idx_cv

    for i in range(2, last_row_b + 1):
        cv_sheet.cell(row=i, column=1).value = i - 1
        cv_sheet.cell(row=i, column=1).border = thin_border

    merged_cells_to_unmerge = []
    for cell_range_str in list(ws.merged_cells.ranges):
        # Corrected usage of openpyxl.utils.cell.range_boundaries
        min_col_mc, min_row_mc, max_col_mc, max_row_mc = openpyxl.utils.cell.range_boundaries(str(cell_range_str))
        if (
           (min_row_mc <= YEAR_ROW <= max_row_mc) or            (min_row_mc <= MONTH_ROW <= max_row_mc) or            (min_row_mc <= START_ROW_DATA <= max_row_mc) or            (min_row_mc <= YEARLY_TOTAL_ROW <= max_row_mc)
        ):
            merged_cells_to_unmerge.append(cell_range_str)

    for cell_range_str in merged_cells_to_unmerge:
        ws.unmerge_cells(str(cell_range_str))

    max_col_to_clear = max(START_COL + len(years) * 12, ws.max_column + 1)

    rows_to_clear_completely = [YEAR_ROW, MONTH_ROW, START_ROW_DATA, YEARLY_TOTAL_ROW]

    for r_clear in rows_to_clear_completely:
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    for r_clear in range(START_ROW_DATA + 2, ws.max_row + 1):
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear,c_clear).value = None
            ws.cell(r_clear,c_clear).fill = PatternFill()

    # ------------------------------------------------
    # Initialize Allocation Data Structures
    # ------------------------------------------------
    yearly_am_totals = {year: 0 for year in years}
    month_allocation_status = {(y, m): None for y in years for m in range(1, 13)}

    # ------------------------------------------------
    # Build Years and Months Headers on 'ΑΝΑΛΥΣΗ' Sheet
    # ------------------------------------------------
    col = START_COL
    month_col_map = {}

    for y in years:
        year_start_col = col
        year_header_cell = ws.cell(YEAR_ROW, col)

        r_color_func = lambda: random.randint(0,255)
        random_color_hex = '%02X%02X%02X' % (r_color_func(), r_color_func(), r_color_func())
        year_header_cell.fill = PatternFill(start_color=random_color_hex, end_color=random_color_hex, fill_type="solid")

        if not is_light_color(random_color_hex):
            year_header_cell.font = Font(color="FFFFFF")
        else:
            year_header_cell.font = Font(color="000000")

        for m in range(1,13):
            ws.cell(MONTH_ROW, col).value = m
            month_col_map[(y,m)] = col
            col += 1
        year_end_col = col - 1

        ws.merge_cells(start_row=YEAR_ROW, start_column=year_start_col, end_row=YEAR_ROW, end_column=year_end_col)
        year_header_cell.value = y

    for c_border in range(START_COL, col):
        ws.cell(YEAR_ROW, c_border).border = thin_border
        ws.cell(MONTH_ROW, c_border).border = thin_border

    ws.cell(YEARLY_TOTAL_ROW, 2).value = "ΕΤΗΣΙΑ ΣΥΝΟΛΑ"
    ws.cell(YEARLY_TOTAL_ROW, 2).font = Font(bold=True)
    ws.cell(YEARLY_TOTAL_ROW, 2).border = thin_border

    # ------------------------------------------------
    # Set 'ΑΝΑΛΥΣΗ' Sheet Specific Headers and Styling
    # ------------------------------------------------
    ws['A5'] = 'Α/Α'
    ws['B2'] = 'ΑΝΘΡΩΠΟΜΗΝΕΣ ΕΜΠΕΙΡΙΑΣ'
    ws['B2'].fill = orange_fill

    # ------------------------------------------------
    # Greedy Allocation - Pass 1
    # ------------------------------------------------
    current_excel_row = START_ROW_DATA + 2

    yearly_overages = {}

    for project_idx, project_data in enumerate(data):
        project_data["excel_row"] = current_excel_row

        period_str = project_data["period_str"]
        original_am = project_data["original_am"]
        months_in_period = project_data["months_in_period"]
        project_id = project_data["project_id"]
        is_yellow_project = project_data["is_yellow"]
        allocated_count = 0
        unallocated_count = original_am

        ws.cell(current_excel_row,1).value = f'=MATCH(B{current_excel_row},CV!$B$2:$B${last_row_b},0)'
        ws.cell(current_excel_row,1).border = thin_border

        ws.cell(current_excel_row,2).value = period_str
        ws.cell(current_excel_row,2).border = thin_border
        ws.cell(current_excel_row,3).value = original_am
        ws.cell(current_excel_row,3).border = thin_border

        if is_yellow_project:
            ws.cell(current_excel_row,2).fill = yellow

        for (y, m) in sorted(months_in_period):
            if allocated_count >= original_am:
                break

            if (y,m) in month_col_map:
                if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                    reason_text = f"Year {y} capacity reached"
                    if reason_text not in project_data["reasons_log"]:
                        project_data["reasons_log"].append(reason_text)
                    continue

                if month_allocation_status[(y,m)] is not None:
                    occupying_project_id = month_allocation_status[(y,m)]
                    reason_text = f"Month {m}/{y} already allocated by Project {occupying_project_id}"
                    if reason_text not in project_data["reasons_log"]:
                        project_data["reasons_log"].append(reason_text)
                    continue

                cell_to_fill = ws.cell(current_excel_row, month_col_map[(y,m)])
                cell_to_fill.value = 'X'
                cell_to_fill.fill = yellow
                cell_to_fill.border = thin_border

                yearly_am_totals[y] += 1
                month_allocation_status[(y,m)] = project_id
                allocated_count += 1
                unallocated_count -= 1

        project_data["allocated_am"] = allocated_count
        project_data["unallocated_am"] = unallocated_count

        for c_border in range(START_COL, col):
            ws.cell(current_excel_row, c_border).border = thin_border

        current_excel_row += 1

    # ------------------------------------------------
    # Greedy Allocation - Pass 2: Enforcement for zero-allocated projects
    # ------------------------------------------------

    projects_that_got_allocated_in_pass2 = []

    for project_data in data:
        if project_data["allocated_am"] == 0 and project_data["original_am"] > 0:
            current_project_id = project_data["project_id"]
            current_project_excel_row = project_data["excel_row"]
            found_allocation_in_pass2 = False
            project_data["reasons_log"] = []

            for (y, m) in sorted(project_data["months_in_period"]):
                if (y,m) in month_col_map and month_allocation_status[(y,m)] is None:
                    if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                        reason_text = f"Year {y} capacity reached (Pass 2, Attempt 1)"
                        if reason_text not in project_data["reasons_log"]:
                            project_data["reasons_log"].append(reason_text)
                        continue

                    month_allocation_status[(y,m)] = current_project_id
                    yearly_am_totals[y] += 1
                    project_data["allocated_am"] += 1
                    project_data["unallocated_am"] -= 1

                    cell_to_fill = ws.cell(current_project_excel_row, month_col_map[(y,m)])
                    cell_to_fill.value = 'X'
                    cell_to_fill.fill = yellow
                    cell_to_fill.border = thin_border

                    found_allocation_in_pass2 = True
                    projects_that_got_allocated_in_pass2.append(current_project_id)
                    break

            if found_allocation_in_pass2:
                continue

            for (y, m) in sorted(project_data["months_in_period"]):
                if (y,m) in month_col_map:
                    occupying_project_id = month_allocation_status[(y,m)]
                    if occupying_project_id is not None and occupying_project_id != current_project_id:
                        donor_project = project_id_map.get(occupying_project_id)
                        if donor_project and not donor_project["is_yellow"] and donor_project["allocated_am"] > 1:
                            donor_project["allocated_am"] -= 1
                            donor_project["unallocated_am"] += 1
                            ws.cell(donor_project["excel_row"], month_col_map[(y,m)]).value = None
                            ws.cell(donor_project["excel_row"], month_col_map[(y,m)]).fill = PatternFill()

                            month_allocation_status[(y,m)] = current_project_id
                            project_data["allocated_am"] += 1
                            project_data["unallocated_am"] -= 1

                            cell_to_fill = ws.cell(current_project_excel_row, month_col_map[(y,m)])
                            cell_to_fill.value = 'X'
                            cell_to_fill.fill = yellow
                            cell_to_fill.border = thin_border

                            found_allocation_in_pass2 = True
                            projects_that_got_allocated_in_pass2.append(current_project_id)
                            break

            if found_allocation_in_pass2:
                continue

            for (y, m) in sorted(project_data["months_in_period"]):
                if (y,m) in month_col_map:
                    occupying_project_id = month_allocation_status[(y,m)]
                    if occupying_project_id is not None and occupying_project_id != current_project_id:
                        donor_project = project_id_map.get(occupying_project_id)
                        if donor_project and donor_project["allocated_am"] > 1:
                            donor_project["allocated_am"] -= 1
                            donor_project["unallocated_am"] += 1
                            ws.cell(donor_project["excel_row"], month_col_map[(y,m)]).value = None
                            ws.cell(donor_project["excel_row"], month_col_map[(y,m)]).fill = PatternFill()

                            month_allocation_status[(y,m)] = current_project_id
                            project_data["allocated_am"] += 1
                            project_data["unallocated_am"] -= 1

                            cell_to_fill = ws.cell(current_project_excel_row, month_col_map[(y,m)])
                            cell_to_fill.value = 'X'
                            cell_to_fill.fill = yellow
                            cell_to_fill.border = thin_border

                            found_allocation_in_pass2 = True
                            projects_that_got_allocated_in_pass2.append(current_project_id)
                            break

    # ------------------------------------------------
    # Post-processing: Enforce strict yearly capacity (11 person-months)
    # ------------------------------------------------
    for y in years:
        while yearly_am_totals[y] > MAX_YEARLY_CAPACITY:
            deallocated_a_month_in_this_iteration = False

            allocated_months_in_year = []
            for m_idx in range(1, 13):
                month_key = (y, m_idx)
                if month_key in month_col_map and month_allocation_status[month_key] is not None:
                    occupying_project_id = month_allocation_status[month_key]
                    project_info = project_id_map.get(occupying_project_id)
                    if project_info:
                        allocated_months_in_year.append((month_key, project_info))

            month_to_deallocate_key = None
            donor_project_id_for_deallocation = None

            for month_key, project_info in allocated_months_in_year:
                if not project_info["is_yellow"] and                    project_info["allocated_am"] > 1 and                    project_info["project_id"] not in projects_that_got_allocated_in_pass2:
                    month_to_deallocate_key = month_key
                    donor_project_id_for_deallocation = project_info["project_id"]
                    break

            if month_to_deallocate_key is None:
                for month_key, project_info in allocated_months_in_year:
                    if not project_info["is_yellow"] and project_info["allocated_am"] > 1:
                        month_to_deallocate_key = month_key
                        donor_project_id_for_deallocation = project_info["project_id"]
                        break

            if month_to_deallocate_key is None:
                for month_key, project_info in allocated_months_in_year:
                    if project_info["allocated_am"] > 1:
                        month_to_deallocate_key = month_key
                        donor_project_id_for_deallocation = project_info["project_id"]
                        break

            if month_to_deallocate_key is None:
                for month_key, project_info in allocated_months_in_year:
                    month_to_deallocate_key = month_key
                    donor_project_id_for_deallocation = project_info["project_id"]
                    break


            if month_to_deallocate_key is not None and donor_project_id_for_deallocation is not None:
                donor_project = project_id_map.get(donor_project_id_for_deallocation)
                if donor_project:
                    month_allocation_status[month_to_deallocate_key] = None
                    yearly_am_totals[y] -= 1
                    donor_project["allocated_am"] -= 1
                    donor_project["unallocated_am"] += 1

                    ws.cell(donor_project["excel_row"], month_col_map[month_to_deallocate_key]).value = None
                    ws.cell(donor_project["excel_row"], month_col_map[month_to_deallocate_key]).fill = PatternFill()

                    reason_text = f"Month {month_to_deallocate_key[1]}/{month_to_deallocate_key[0]} deallocated due to year {y} capacity enforcement."
                    if reason_text not in donor_project["reasons_log"]:
                        donor_project["reasons_log"].append(reason_text)

                    deallocated_a_month_in_this_iteration = True

            if deallocated_a_month_in_this_iteration:
                continue
            else:
                st.warning(f"Warning: No suitable month could be deallocated in year {y} to meet capacity (total: {yearly_am_totals[y]}). All remaining allocations might be protected by rules.")
                break

    total_yellow_allocated_am_final = 0
    for project_data in data:
        if project_data["is_yellow"]:
            total_yellow_allocated_am_final += project_data["allocated_am"]

    unallocated_projects = []
    for project_data in data:
        if project_data["unallocated_am"] > 0:
            ws.cell(project_data["excel_row"], 3).font = Font(color="FF0000", bold=True)

            final_reasons = "; ".join(list(set(project_data["reasons_log"]))) if project_data["reasons_log"] else "Capacity/Month taken by other projects."

            unallocated_projects.append({
                "period": project_data["period_str"],
                "original_am": project_data["original_am"],
                "allocated_am": project_data["allocated_am"],
                "unallocated_am": project_data["unallocated_am"],
                "reasons": final_reasons
            })
        else:
            ws.cell(project_data["excel_row"], 3).font = Font(color="000000")


    # ------------------------------------------------
    # Finalize 'ΑΝΑΛΥΣΗ' Sheet Totals and Styling
    # ------------------------------------------------

    for y in years:
        if y in yearly_am_totals:
            col_for_year_total = month_col_map[(y, 1)]
            year_month_cols = [month_col_map[(y, m)] for m in range(1, 13) if (y, m) in month_col_map]
            if year_month_cols:
                year_start_col = min(year_month_cols)
                year_end_col = max(year_month_cols)
                if year_start_col != year_end_col:
                    ws.merge_cells(start_row=YEARLY_TOTAL_ROW, start_column=year_start_col, end_row=YEARLY_TOTAL_ROW, end_column=year_end_col)
                col_for_year_total = year_start_col
            else:
                continue

            total_cell = ws.cell(YEARLY_TOTAL_ROW, col_for_year_total)
            total_cell.value = yearly_am_totals[y]
            total_cell.font = Font(bold=True)
            total_cell.border = thin_border
            total_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                year_header_cell = ws.cell(YEAR_ROW, col_for_year_total)
                year_header_cell.fill = red_fill
                year_header_cell.font = Font(color="FFFFFF", bold=True)

                total_cell.fill = red_fill
                total_cell.font = Font(color="FFFFFF", bold=True)
                if yearly_am_totals[y] > MAX_YEARLY_CAPACITY:
                    yearly_overages[y] = yearly_am_totals[y] - MAX_YEARLY_CAPACITY
            elif yearly_am_totals[y] > 0:
                total_cell.fill = green_fill
                total_cell.font = Font(color="000000", bold=True)

    if total_yellow_allocated_am_final > 0:
        ws['C2'].value = total_yellow_allocated_am_final
    else:
        ws['C2'].value = total_all_projects_am
    ws['C2'].font = Font(bold=True)
    ws['C2'].border = thin_border
    ws['C2'].fill = orange_fill

    for c_width in range(START_COL, col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_width)].width = 2.5

    ws['A6'] = f'=MATCH(B6,CV!$B$2:$B${last_row_b},0)'

    # Save the workbook to a BytesIO object
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer, unallocated_projects, yearly_am_totals, yearly_overages, MAX_YEARLY_CAPACITY


st.set_page_config(layout="wide", page_title="Ανθρωπομήνες - Κατανομή Έργων")

# Define the logo URL
LOGO_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-7/main/SPACE%20LOGO_colored%20horizontal.png"

# Create two columns for the title and the logo
col1, col2 = st.columns([3, 1]) # Adjust ratio as needed

with col1:
    st.title("Εργαλείο Κατανομής Ανθρωπομηνών")
    st.markdown("Αυτό το εργαλείο κατανέμει ανθρωπομήνες σε έργα με βάση χρονικά διαστήματα και μέγιστη ετήσια χωρητικότητα.")

with col2:
    st.image(LOGO_URL, width=378) # Adjusted width to 10cm (approx 378px)

# File Uploaders
template_file = st.file_uploader("Παρακαλώ ανεβάστε το αρχείο template (οποιοδήποτε excel) - (AM TEST 1.xlsx)", type=["xlsx"])
input_file = st.file_uploader("👉 Ανέβασε το INPUT excel (μόνο 2 στήλες)", type=["xlsx"]) # Updated input uploader label

template_file_bytes = None # Initialize

# Handle template file upload or default loading
if template_file is None:
    st.info("Trying to load default template 'AM TEST 1.xlsx'...")
    st.info(f"Current working directory: {os.getcwd()}")
    st.info(f"Does 'AM TEST 1.xlsx' exist? {os.path.exists('AM TEST 1.xlsx')}")
    try:
        # Assuming 'AM TEST 1.xlsx' is in the same directory as app.py
        with open("AM TEST 1.xlsx", "rb") as f:
            template_file_bytes = io.BytesIO(f.read())
        st.info("Χρησιμοποιήθηκε το προεπιλεγμένο αρχείο template: 'AM TEST 1.xlsx'")
    except FileNotFoundError:
        st.warning("Το προεπιλεγμένο αρχείο 'AM TEST 1.xlsx' δεν βρέθηκε. Παρακαλώ ανεβάστε ένα αρχείο template ή βεβαιωθείτε ότι το 'AM TEST 1.xlsx' υπάρχει στον ίδιο φάκελο στο GitHub.")
        template_file_bytes = None
else:
    template_file_bytes = io.BytesIO(template_file.read())

# Logic for processing or showing initial instruction
if input_file is None:
    st.info("Παρακαλώ ανεβάστε ένα αρχείο Excel για να ξεκινήσετε την επεξεργασία.")
elif template_file_bytes is None: # Input file is present, but template isn't (e.g., default failed and user didn't upload)
    st.warning("Το αρχείο template δεν είναι διαθέσιμο. Παρακαλώ ανεβάστε το.")
else: # Both input and template are available
    if st.button("Εκτέλεση Κατανομής"): 
        with st.spinner('Επεξεργασία του αρχείου...'):
            output_excel_buffer, unallocated_projects, yearly_am_totals, yearly_overages, MAX_YEARLY_CAPACITY = process_excel_data(template_file_bytes, io.BytesIO(input_file.read()))
            
            if output_excel_buffer:
                st.success("Το αρχείο επεξεργάστηκε επιτυχώς!")
                st.download_button(
                    label="Κατεβάστε το επεξεργασμένο Excel",
                    data=output_excel_buffer,
                    file_name=f"{os.path.splitext(input_file.name)[0]}_ΚΑΤΑΝΟΜΗ ΑΜ.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.subheader("Σύνοψη Κατανομής")
                st.write(f"Μέγιστη ετήσια χωρητικότητα ανά έτος: {MAX_YEARLY_CAPACITY} ανθρωπομήνες")

                st.markdown("**Ετήσια σύνολα ανθρωπομηνών:**")
                for year, total_am in sorted(yearly_am_totals.items()):
                    status = " (Η χωρητικότητα επιτεύχθηκε)" if total_am >= MAX_YEARLY_CAPACITY else ""
                    if year in yearly_overages and yearly_overages[year] > 0:
                        status = f" (ΥΠΕΡΒΑΣΗ ΧΩΡΗΤΙΚΟΤΗΤΑΣ κατά {yearly_overages[year]})"
                    st.write(f"  Έτος {year}: {total_am}{status}")

                if unallocated_projects:
                    st.markdown("**Έργα με μη κατανεμημένους ανθρωπομήνες:**")
                    for proj in unallocated_projects:
                        st.write(f"  Περίοδος: {proj['period']}, Αρχικοί ΑΜ: {proj['original_am']}, Κατανεμημένοι ΑΜ: {proj['allocated_am']}, Μη κατανεμημένοι ΑΜ: {proj['unallocated_am']}")
                        if proj['reasons']:
                            st.markdown(f"    Λόγοι μη κατανομής: _{proj['reasons']}_ ")
                else:
                    st.success("Όλοι οι ανθρωπομήνες κατανεμήθηκαν επιτυχώς.")
            else:
                st.error("Παρουσιάστηκε σφάλμα κατά την επεξεργασία του αρχείου. Παρακαλώ ελέγξτε τα αρχεία εισόδου.")

